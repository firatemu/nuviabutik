from decimal import Decimal
import json

from django.shortcuts import render, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.http import HttpResponse, JsonResponse
from django.views.decorators.http import require_POST
from django.db.models import Sum, Count, F, Max, ExpressionWrapper, DecimalField
from datetime import date, datetime, timedelta
from openpyxl import Workbook
from reportlab.pdfgen import canvas
from satis.models import Satis, SatisDetay
from urun.models import Urun
from musteri.models import Musteri


@login_required
def gunluk_satis(request):
    """Günlük satış raporu view'ı"""
    bugun = date.today()
    tarih = request.GET.get('tarih', bugun.strftime('%Y-%m-%d'))

    try:
        secili_tarih = datetime.strptime(tarih, '%Y-%m-%d').date()
    except ValueError:
        secili_tarih = bugun

    # Günlük satışlar - detaylı bilgi ile
    satislar = Satis.objects.filter(
        satis_tarihi__date=secili_tarih,
        durum='tamamlandi'
    ).select_related('musteri', 'satici').prefetch_related('satisdetay_set__varyant__urun__kategori', 'satisdetay_set__varyant__urun__marka')

    # Günlük satış detayları - ürün bazında
    satis_detaylari = SatisDetay.objects.filter(
        satis__satis_tarihi__date=secili_tarih,
        satis__durum='tamamlandi'
    ).select_related(
        'satis', 'satis__musteri', 'satis__satici',
        'urun', 'urun__kategori', 'urun__marka',
        'varyant', 'varyant__urun',
        'varyant__urun__kategori', 'varyant__urun__marka',
        'varyant__renk', 'varyant__beden'
    ).order_by('-satis__satis_tarihi')

    # İstatistikler
    toplam_satis = satislar.aggregate(
        toplam=Sum('toplam_tutar'),
        adet=Count('id')
    )

    toplam_urun_sayisi = satis_detaylari.aggregate(
        toplam_adet=Sum('miktar')
    )['toplam_adet'] or 0

    # Ortalama satış hesapla
    ortalama_satis = 0
    if toplam_satis['adet'] and toplam_satis['adet'] > 0:
        ortalama_satis = toplam_satis['toplam'] / toplam_satis['adet']

    context = {
        'satislar': satislar,
        'satis_detaylari': satis_detaylari,
        'tarih': secili_tarih.strftime('%Y-%m-%d'),
        'toplam_satis': toplam_satis['toplam'] or 0,
        'satis_sayisi': toplam_satis['adet'] or 0,
        'toplam_urun_sayisi': toplam_urun_sayisi,
        'ortalama_satis': ortalama_satis,
        'toplam_urun_adet': toplam_urun_sayisi,  # Template'te bu isim kullanılıyor
    }
    return render(request, 'rapor/gunluk_satis.html', context)


@login_required
def stok_raporu(request):
    """Stok raporu view'ı"""
    from urun.models import UrunVaryanti, UrunKategoriUst, Marka
    from django.db.models import Sum, Q

    # Tüm varyantları al ve ürün bazında grupla
    varyantlar = UrunVaryanti.objects.filter(
        aktif=True,
        urun__aktif=True
    ).select_related('urun', 'urun__kategori', 'urun__marka', 'renk', 'beden').order_by('urun__kategori__ad', 'urun__ad')

    # Arama filtreleri
    arama = request.GET.get('arama', '').strip()
    kategori_id = request.GET.get('kategori')
    marka_id = request.GET.get('marka')
    durum = request.GET.get('durum')
    cinsiyet = request.GET.get('cinsiyet')
    kar_orani_min = request.GET.get('kar_orani_min', '').strip()
    kar_orani_max = request.GET.get('kar_orani_max', '').strip()

    # Sıralama parametreleri
    sort_field = request.GET.get('sort', 'urun__ad')
    sort_order = request.GET.get('order', 'asc')

    # Geçerli sıralama alanları
    valid_sort_fields = [
        'urun__ad', 'renk__ad', 'barkod', 'urun__kategori__ad',
        'urun__marka__ad', 'urun__alis_fiyati', 'urun__satis_fiyati', 'urun__kar_orani', 'stok_miktari'
    ]

    if sort_field not in valid_sort_fields:
        sort_field = 'urun__ad'

    # Sıralama yönü
    if sort_order == 'desc':
        sort_field = '-' + sort_field

    # Arama filtresi
    if arama:
        varyantlar = varyantlar.filter(
            Q(urun__ad__icontains=arama) |
            Q(barkod__icontains=arama) |
            Q(urun__urun_kodu__icontains=arama) |
            Q(renk__ad__icontains=arama) |
            Q(beden__ad__icontains=arama)
        )

    # Kategori filtresi
    if kategori_id:
        varyantlar = varyantlar.filter(urun__kategori_id=kategori_id)

    # Marka filtresi
    if marka_id:
        varyantlar = varyantlar.filter(urun__marka_id=marka_id)

    # Stok durumu filtresi - Yeni 3 seçenek: Hepsi / Stoğu Olan / Stoğu Biten
    if durum == 'stogu_olan':
        varyantlar = varyantlar.filter(stok_miktari__gt=0)
    elif durum == 'stogu_biten':
        varyantlar = varyantlar.filter(stok_miktari=0)
    # '' (boş) veya yok → tümü

    # Cinsiyet filtresi
    if cinsiyet and cinsiyet != 'hepsi':
        varyantlar = varyantlar.filter(urun__cinsiyet=cinsiyet)

    # Sıralama uygula
    varyantlar = varyantlar.order_by(
        sort_field, 'urun__ad', 'renk__ad', 'beden__ad')

    # Filtrelenmiş varyantları listeye çevir (Python tarafı kar oranı hesaplaması için)
    varyant_list = list(varyantlar)

    # Kar oranı aralık filtresi - Python tarafında uygulanır
    min_val = None
    max_val = None
    try:
        if kar_orani_min:
            min_val = float(kar_orani_min)
    except (ValueError, TypeError):
        min_val = None
    try:
        if kar_orani_max:
            max_val = float(kar_orani_max)
    except (ValueError, TypeError):
        max_val = None

    def _passes_kar(v):
        alis = v.urun.alis_fiyati or Decimal('0')
        satis = v.urun.pesin_fiyat or Decimal('0')
        if satis != 0:
            oran = float((satis - alis) / satis * Decimal('100'))
        else:
            oran = 0.0
        if min_val is not None and oran < min_val:
            return False
        if max_val is not None and oran > max_val:
            return False
        return True

    if min_val is not None or max_val is not None:
        varyant_list = [v for v in varyant_list if _passes_kar(v)]

    # Her varyant için kar tutarı ve kar oranı hesapla (template'e hazır dict)
    ZERO = Decimal('0')
    rows = []
    for v in varyant_list:
        alis = v.urun.alis_fiyati or ZERO
        satis = v.urun.pesin_fiyat or ZERO
        kar_tutari = satis - alis
        if satis != 0:
            kar_orani = (kar_tutari / satis) * Decimal('100')
        else:
            kar_orani = ZERO
        rows.append({
            'varyant': v,
            'kar_tutari': kar_tutari,
            'kar_orani': kar_orani,
        })

    # Dropdown için veriler
    kategoriler = UrunKategoriUst.objects.all().order_by('ad')
    markalar = Marka.objects.all().order_by('ad')

    context = {
        'rows': rows,
        'kategoriler': kategoriler,
        'markalar': markalar,
        'arama': arama,
        'kategori_id': kategori_id,
        'marka_id': marka_id,
        'durum': durum,
        'cinsiyet': cinsiyet,
        'kar_orani_min': kar_orani_min,
        'kar_orani_max': kar_orani_max,
        'sort_field': request.GET.get('sort', 'urun__ad'),
        'sort_order': request.GET.get('order', 'asc'),
    }
    return render(request, 'rapor/stok_raporu.html', context)


def _stok_degeri_by_urun():
    """Aktif varyantlardan ürün başına stok miktarı ve ürün nesnesi."""
    from collections import defaultdict
    from urun.models import UrunVaryanti

    by_urun = defaultdict(int)
    urun_map = {}
    varyantlar = UrunVaryanti.objects.filter(
        aktif=True,
        urun__aktif=True,
        stok_miktari__gt=0,
    ).select_related('urun', 'urun__kategori', 'urun__marka')
    for v in varyantlar:
        by_urun[v.urun_id] += v.stok_miktari
        urun_map[v.urun_id] = v.urun
    return by_urun, urun_map


def _urun_stok_toplami(urun_id, by_urun=None):
    if by_urun is None:
        by_urun, _ = _stok_degeri_by_urun()
    return by_urun.get(urun_id, 0)


def _stok_degeri_filter_params(request):
    """GET filtreleri: kategori, marka, alis_fiyat ('', sifir, sifir_degil)."""
    kategori = (request.GET.get('kategori') or '').strip()
    marka = (request.GET.get('marka') or '').strip()
    alis_fiyat = (request.GET.get('alis_fiyat') or '').strip()
    if alis_fiyat not in ('', 'sifir', 'sifir_degil'):
        alis_fiyat = ''

    kategori_id = None
    if kategori:
        try:
            kategori_id = int(kategori)
        except (TypeError, ValueError):
            kategori = ''
            kategori_id = None

    marka_id = None
    if marka:
        try:
            marka_id = int(marka)
        except (TypeError, ValueError):
            marka = ''
            marka_id = None

    return {
        'kategori': kategori,
        'marka': marka,
        'alis_fiyat': alis_fiyat,
        'kategori_id': kategori_id,
        'marka_id': marka_id,
    }


def _urun_matches_stok_degeri_filters(urun, alis, kategori_id=None, marka_id=None, alis_fiyat=''):
    if kategori_id is not None and urun.kategori_id != kategori_id:
        return False
    if marka_id is not None and urun.marka_id != marka_id:
        return False
    if alis_fiyat == 'sifir' and alis != 0:
        return False
    if alis_fiyat == 'sifir_degil' and alis <= 0:
        return False
    return True


def _build_stok_degeri_liste(
    by_urun=None,
    urun_map=None,
    kategori_id=None,
    marka_id=None,
    alis_fiyat='',
):
    if by_urun is None or urun_map is None:
        by_urun, urun_map = _stok_degeri_by_urun()

    urun_stok_degerleri = []
    toplam_stok_degeri = Decimal('0')
    for uid, ts in by_urun.items():
        urun = urun_map[uid]
        alis = urun.alis_fiyati or Decimal('0')
        if not isinstance(alis, Decimal):
            alis = Decimal(str(alis))
        if not _urun_matches_stok_degeri_filters(
            urun, alis, kategori_id=kategori_id, marka_id=marka_id, alis_fiyat=alis_fiyat
        ):
            continue
        deger = alis * ts
        toplam_stok_degeri += deger
        urun_stok_degerleri.append({
            'urun': urun,
            'alis_fiyati': alis,
            'toplam_stok': ts,
            'stok_degeri': deger,
        })
    urun_stok_degerleri.sort(key=lambda x: x['stok_degeri'], reverse=True)
    return urun_stok_degerleri, toplam_stok_degeri


@login_required
def stok_degeri(request):
    """Ürün bazında stok maliyeti (alış × stok)."""
    from urun.models import UrunKategoriUst, Marka

    filters = _stok_degeri_filter_params(request)
    by_urun, urun_map = _stok_degeri_by_urun()
    urun_stok_degerleri, toplam_stok_degeri = _build_stok_degeri_liste(
        by_urun,
        urun_map,
        kategori_id=filters['kategori_id'],
        marka_id=filters['marka_id'],
        alis_fiyat=filters['alis_fiyat'],
    )

    context = {
        'toplam_urun_sayisi': len(urun_stok_degerleri),
        'toplam_stok_degeri': toplam_stok_degeri,
        'urun_stok_degerleri': urun_stok_degerleri,
        'kategori': filters['kategori'],
        'marka': filters['marka'],
        'alis_fiyat': filters['alis_fiyat'],
        'kategori_listesi': UrunKategoriUst.objects.filter(aktif=True).order_by('ad'),
        'marka_listesi': Marka.objects.filter(aktif=True).order_by('ad'),
    }
    return render(request, 'rapor/stok_degeri.html', context)


@login_required
@require_POST
def stok_degeri_alis_fiyat_guncelle(request):
    """Yalnızca ürün alış fiyatını günceller (save/sinyal tetiklenmez)."""
    if request.content_type and 'application/json' in request.content_type:
        try:
            payload = json.loads(request.body)
        except json.JSONDecodeError:
            return JsonResponse({'success': False, 'message': 'Geçersiz JSON.'}, status=400)
        urun_id = payload.get('urun_id')
        alis_raw = payload.get('alis_fiyati')
    else:
        urun_id = request.POST.get('urun_id')
        alis_raw = request.POST.get('alis_fiyati')

    try:
        urun_id = int(urun_id)
    except (TypeError, ValueError):
        return JsonResponse({'success': False, 'message': 'Geçersiz ürün.'}, status=400)

    try:
        alis = Decimal(str(alis_raw).replace(',', '.'))
    except Exception:
        return JsonResponse({'success': False, 'message': 'Geçersiz alış fiyatı.'}, status=400)

    if alis < 0:
        return JsonResponse({'success': False, 'message': 'Alış fiyatı negatif olamaz.'}, status=400)

    urun = Urun.objects.filter(pk=urun_id, aktif=True).first()
    if not urun:
        return JsonResponse({'success': False, 'message': 'Ürün bulunamadı.'}, status=404)

    toplam_stok = _urun_stok_toplami(urun_id)
    if toplam_stok <= 0:
        return JsonResponse(
            {'success': False, 'message': 'Bu ürünün stoğu olan kaydı yok.'},
            status=400,
        )

    updated = Urun.objects.filter(pk=urun_id, aktif=True).update(alis_fiyati=alis)
    if not updated:
        return JsonResponse({'success': False, 'message': 'Güncelleme yapılamadı.'}, status=400)

    filters = _stok_degeri_filter_params(request)
    by_urun, urun_map = _stok_degeri_by_urun()
    _, toplam_stok_degeri = _build_stok_degeri_liste(
        by_urun,
        urun_map,
        kategori_id=filters['kategori_id'],
        marka_id=filters['marka_id'],
        alis_fiyat=filters['alis_fiyat'],
    )
    stok_degeri = alis * toplam_stok

    return JsonResponse({
        'success': True,
        'alis_fiyati': f'{alis:.2f}',
        'stok_degeri': f'{stok_degeri:.2f}',
        'toplam_stok_degeri': f'{toplam_stok_degeri:.2f}',
    })


def _fatura_karlilik_tarih_araligi(request):
    baslangic = request.GET.get('baslangic')
    bitis = request.GET.get('bitis')
    if not baslangic:
        baslangic = date.today().replace(day=1)
    else:
        baslangic = datetime.strptime(baslangic, '%Y-%m-%d').date()
    if not bitis:
        bitis = date.today()
    else:
        bitis = datetime.strptime(bitis, '%Y-%m-%d').date()
    if baslangic > bitis:
        baslangic, bitis = bitis, baslangic
    return baslangic, bitis


def _fatura_geneli_kar_ve_ciro_marji(fatura_net, maliyet):
    fn = fatura_net or Decimal('0')
    mal = maliyet or Decimal('0')
    kar = fn - mal
    marj = (kar / fn * Decimal('100')) if fn > 0 else Decimal('0')
    return kar, marj


def _kar_yuzde_maliyet_ustu(kar, maliyet):
    mal = maliyet or Decimal('0')
    k = kar or Decimal('0')
    return (k / mal * Decimal('100')) if mal > 0 else Decimal('0')


@login_required
def urun_bazli_karlilik(request):
    """Tamamlanan satışlar üzerinden ürün bazlı karlılık."""
    baslangic = request.GET.get('baslangic')
    bitis = request.GET.get('bitis')
    if not baslangic:
        baslangic = date.today().replace(day=1)
    else:
        baslangic = datetime.strptime(baslangic, '%Y-%m-%d').date()
    if not bitis:
        bitis = date.today()
    else:
        bitis = datetime.strptime(bitis, '%Y-%m-%d').date()
    if baslangic > bitis:
        baslangic, bitis = bitis, baslangic

    money = DecimalField(max_digits=14, decimal_places=2)
    qs = SatisDetay.objects.filter(
        satis__satis_tarihi__date__range=[baslangic, bitis],
        satis__durum='tamamlandi',
    )
    agg = qs.values('urun_id', 'urun__ad', 'urun__urun_kodu').annotate(
        toplam_miktar=Sum('miktar'),
        birim_alis=Max('urun__alis_fiyati'),
        brut_tutar=Sum(
            ExpressionWrapper(
                F('birim_fiyat') * F('miktar'),
                output_field=money,
            )
        ),
        toplam_indirim=Sum('indirim_tutari'),
        net_satis=Sum('toplam_fiyat'),
        toplam_maliyet=Sum(
            ExpressionWrapper(
                F('urun__alis_fiyati') * F('miktar'),
                output_field=money,
            )
        ),
    ).order_by('-net_satis')

    satirlar = []
    for row in agg:
        net = row['net_satis'] or Decimal('0')
        mal = row['toplam_maliyet'] or Decimal('0')
        kar = net - mal
        kar_orani = (kar / mal * Decimal('100')) if mal and mal > 0 else Decimal('0')
        satirlar.append({**row, 'kar': kar, 'kar_orani': kar_orani})

    toplam_miktar = sum((r['toplam_miktar'] or 0) for r in satirlar)
    toplam_brut = sum((r['brut_tutar'] or Decimal('0')) for r in satirlar)
    toplam_indirim = sum((r['toplam_indirim'] or Decimal('0')) for r in satirlar)
    toplam_net = sum((r['net_satis'] or Decimal('0')) for r in satirlar)
    toplam_maliyet = sum((r['toplam_maliyet'] or Decimal('0')) for r in satirlar)
    toplam_kar = toplam_net - toplam_maliyet
    genel_kar_orani = (
        (toplam_kar / toplam_maliyet * Decimal('100'))
        if toplam_maliyet and toplam_maliyet > 0
        else Decimal('0')
    )

    context = {
        'title': 'Ürün Bazlı Karlılık',
        'baslangic': baslangic,
        'bitis': bitis,
        'satirlar': satirlar,
        'ozet': {
            'urun_sayisi': len(satirlar),
            'toplam_miktar': toplam_miktar,
            'toplam_brut': toplam_brut,
            'toplam_indirim': toplam_indirim,
            'toplam_net': toplam_net,
            'toplam_maliyet': toplam_maliyet,
            'toplam_kar': toplam_kar,
            'genel_kar_orani': genel_kar_orani,
        },
    }
    return render(request, 'rapor/urun_bazli_karlilik.html', context)


@login_required
def fatura_bazli_karlilik(request):
    """Tamamlanan satışlar (fatura) bazında karlılık."""
    baslangic, bitis = _fatura_karlilik_tarih_araligi(request)
    money = DecimalField(max_digits=14, decimal_places=2)
    brut_expr = ExpressionWrapper(
        F('satisdetay__birim_fiyat') * F('satisdetay__miktar'),
        output_field=money,
    )
    maliyet_expr = ExpressionWrapper(
        F('satisdetay__urun__alis_fiyati') * F('satisdetay__miktar'),
        output_field=money,
    )
    qs = (
        Satis.objects.filter(
            satis_tarihi__date__range=[baslangic, bitis],
            durum='tamamlandi',
        )
        .select_related('musteri', 'satici')
        .annotate(
            krl_brut_satir=Sum(brut_expr),
            krl_indirim_satir=Sum('satisdetay__indirim_tutari'),
            krl_net_satir=Sum('satisdetay__toplam_fiyat'),
            krl_maliyet=Sum(maliyet_expr),
        )
        .order_by('-satis_tarihi', '-id')
    )

    satirlar = []
    for s in qs:
        fatura_net = s.genel_toplam or Decimal('0')
        mal = s.krl_maliyet or Decimal('0')
        kar, _ = _fatura_geneli_kar_ve_ciro_marji(fatura_net, mal)
        kar_orani = _kar_yuzde_maliyet_ustu(kar, mal)
        satirlar.append({
            'satis': s,
            'brut': s.krl_brut_satir or Decimal('0'),
            'indirim': s.krl_indirim_satir or Decimal('0'),
            'net_satir': s.krl_net_satir or Decimal('0'),
            'fatura_net': fatura_net,
            'maliyet': mal,
            'kar': kar,
            'kar_orani': kar_orani,
        })

    toplam_fatura_net = sum((r['fatura_net'] for r in satirlar), Decimal('0'))
    toplam_maliyet = sum((r['maliyet'] for r in satirlar), Decimal('0'))
    toplam_kar, _ = _fatura_geneli_kar_ve_ciro_marji(toplam_fatura_net, toplam_maliyet)
    genel_kar_orani = _kar_yuzde_maliyet_ustu(toplam_kar, toplam_maliyet)
    toplam_brut = sum((r['brut'] for r in satirlar), Decimal('0'))
    toplam_indirim = sum((r['indirim'] for r in satirlar), Decimal('0'))

    context = {
        'title': 'Fatura Bazlı Karlılık',
        'baslangic': baslangic,
        'bitis': bitis,
        'satirlar': satirlar,
        'ozet': {
            'fatura_adedi': len(satirlar),
            'toplam_brut': toplam_brut,
            'toplam_indirim': toplam_indirim,
            'toplam_fatura_net': toplam_fatura_net,
            'toplam_maliyet': toplam_maliyet,
            'toplam_kar': toplam_kar,
            'genel_kar_orani': genel_kar_orani,
        },
    }
    return render(request, 'rapor/fatura_bazli_karlilik.html', context)


@login_required
def fatura_karlilik_detay(request, pk):
    """Tek satışın satır bazlı karlılık dökümü."""
    baslangic, bitis = _fatura_karlilik_tarih_araligi(request)
    satis = get_object_or_404(Satis, pk=pk, durum='tamamlandi')
    detay_qs = satis.satisdetay_set.select_related(
        'urun', 'varyant', 'varyant__renk', 'varyant__beden'
    ).order_by('id')

    money = DecimalField(max_digits=14, decimal_places=2)
    brut_expr = ExpressionWrapper(
        F('birim_fiyat') * F('miktar'),
        output_field=money,
    )
    agg = detay_qs.aggregate(
        brut=Sum(brut_expr),
        indirim=Sum('indirim_tutari'),
        net_satir=Sum('toplam_fiyat'),
        maliyet=Sum(
            ExpressionWrapper(
                F('urun__alis_fiyati') * F('miktar'),
                output_field=money,
            )
        ),
    )

    detay_satirlari = []
    for d in detay_qs:
        brut = (d.birim_fiyat or Decimal('0')) * Decimal(d.miktar)
        ind = d.indirim_tutari or Decimal('0')
        net = d.toplam_fiyat or Decimal('0')
        alis = d.urun.alis_fiyati or Decimal('0')
        mal = alis * Decimal(d.miktar)
        kar = net - mal
        if mal and mal > 0:
            kar_orani = kar / mal * Decimal('100')
        elif net and net > 0:
            kar_orani = kar / net * Decimal('100')
        else:
            kar_orani = Decimal('0')
        varyant_etiket = ''
        if d.varyant_id:
            parca = []
            if getattr(d.varyant, 'renk_id', None) and d.varyant.renk:
                parca.append(d.varyant.renk.ad)
            if getattr(d.varyant, 'beden_id', None) and d.varyant.beden:
                parca.append(d.varyant.beden.ad)
            varyant_etiket = ' · '.join(parca) if parca else (d.varyant.varyasyon_adi or '')

        detay_satirlari.append({
            'detay': d,
            'brut': brut,
            'indirim': ind,
            'net': net,
            'birim_alis': alis,
            'maliyet': mal,
            'kar': kar,
            'kar_orani': kar_orani,
            'varyant_etiket': varyant_etiket,
        })

    fatura_net = satis.genel_toplam or Decimal('0')
    mal_top = agg['maliyet'] or Decimal('0')
    net_satir = agg['net_satir'] or Decimal('0')
    kar_top, kar_orani_ciro_marji = _fatura_geneli_kar_ve_ciro_marji(fatura_net, mal_top)
    kar_orani_maliyet_top = _kar_yuzde_maliyet_ustu(kar_top, mal_top)
    kar_satir_toplam = net_satir - mal_top
    kar_orani_satir_maliyet = _kar_yuzde_maliyet_ustu(kar_satir_toplam, mal_top)

    context = {
        'title': f"Fatura karlılık — {satis.satis_no or satis.siparis_no}",
        'baslangic': baslangic,
        'bitis': bitis,
        'satis': satis,
        'detay_satirlari': detay_satirlari,
        'ozet': {
            'brut': agg['brut'] or Decimal('0'),
            'indirim': agg['indirim'] or Decimal('0'),
            'net_satir': net_satir,
            'fatura_net': fatura_net,
            'maliyet': mal_top,
            'kar': kar_top,
            'kar_orani': kar_orani_maliyet_top,
            'kar_orani_ciro_marji': kar_orani_ciro_marji,
            'kar_satir_toplam': kar_satir_toplam,
            'kar_orani_satir_maliyet': kar_orani_satir_maliyet,
        },
    }
    return render(request, 'rapor/fatura_karlilik_detay.html', context)


@login_required
def cok_satan_urunler(request):
    """En çok satan ürünler raporu view'ı"""
    # Tarih aralığı
    baslangic = request.GET.get('baslangic')
    bitis = request.GET.get('bitis')

    if not baslangic:
        # Varsayılan: Bu ay
        baslangic = date.today().replace(day=1)
    else:
        baslangic = datetime.strptime(baslangic, '%Y-%m-%d').date()

    if not bitis:
        bitis = date.today()
    else:
        bitis = datetime.strptime(bitis, '%Y-%m-%d').date()

    # En çok satan ürünler
    cok_satanlar = SatisDetay.objects.filter(
        satis__satis_tarihi__date__range=[baslangic, bitis],
        satis__durum='tamamlandi'
    ).values('urun').annotate(
        toplam_miktar=Sum('miktar'),
        toplam_ciro=Sum('toplam_fiyat')
    ).order_by('-toplam_miktar')[:20]

    # Ürün bilgilerini ekle
    for item in cok_satanlar:
        item['urun_obj'] = Urun.objects.get(pk=item['urun'])

    context = {
        'cok_satanlar': cok_satanlar,
        'baslangic': baslangic,
        'bitis': bitis,
    }
    return render(request, 'rapor/cok_satan_urunler.html', context)


@login_required
def kar_zarar(request):
    """Kâr/Zarar analizi view'ı"""
    # Tarih aralığı
    baslangic = request.GET.get('baslangic')
    bitis = request.GET.get('bitis')

    if not baslangic:
        baslangic = date.today().replace(day=1)
    else:
        baslangic = datetime.strptime(baslangic, '%Y-%m-%d').date()

    if not bitis:
        bitis = date.today()
    else:
        bitis = datetime.strptime(bitis, '%Y-%m-%d').date()

    # Satış detayları
    satis_detaylari = SatisDetay.objects.filter(
        satis__satis_tarihi__date__range=[baslangic, bitis],
        satis__durum='tamamlandi'
    )

    # Kâr/Zarar hesaplama
    toplam_ciro = 0
    toplam_maliyet = 0

    for detay in satis_detaylari:
        toplam_ciro += detay.toplam_fiyat
        toplam_maliyet += (detay.urun.alis_fiyati * detay.miktar)

    toplam_kar = toplam_ciro - toplam_maliyet
    kar_marji = (toplam_kar / toplam_ciro * 100) if toplam_ciro > 0 else 0

    context = {
        'toplam_ciro': toplam_ciro,
        'toplam_maliyet': toplam_maliyet,
        'toplam_kar': toplam_kar,
        'kar_marji': kar_marji,
        'baslangic': baslangic,
        'bitis': bitis,
    }
    return render(request, 'rapor/kar_zarar.html', context)


@login_required
def musteri_raporu(request):
    """Müşteri raporu view'ı"""
    musteriler = Musteri.objects.filter(aktif=True)

    # Müşteri satış istatistikleri
    musteri_stats = []
    for musteri in musteriler:
        satislar = Satis.objects.filter(musteri=musteri, durum='tamamlandi')
        toplam_satis = satislar.aggregate(
            toplam=Sum('toplam_tutar'))['toplam'] or 0
        satis_adedi = satislar.count()

        if satis_adedi > 0:
            musteri_stats.append({
                'musteri': musteri,
                'toplam_satis': toplam_satis,
                'satis_adedi': satis_adedi,
                'ortalama_satis': toplam_satis / satis_adedi
            })

    # En çok alışveriş yapan müşteriler
    musteri_stats.sort(key=lambda x: x['toplam_satis'], reverse=True)

    context = {
        'musteri_stats': musteri_stats[:20],
    }
    return render(request, 'rapor/musteri_raporu.html', context)


# Excel Export Views
@login_required
def gunluk_satis_excel(request):
    """Günlük satış Excel export"""
    tarih = request.GET.get('tarih', date.today().strftime('%Y-%m-%d'))
    secili_tarih = datetime.strptime(tarih, '%Y-%m-%d').date()

    satislar = Satis.objects.filter(
        satis_tarihi__date=secili_tarih,
        durum='tamamlandi'
    )

    # Excel dosyası oluştur
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = f"Günlük Satış - {secili_tarih}"

    # Başlıklar
    headers = ['Satış No', 'Müşteri', 'Toplam Tutar', 'Ödeme Tipi', 'Tarih']
    for col, header in enumerate(headers, 1):
        worksheet.cell(row=1, column=col, value=header)

    # Veriler
    for row, satis in enumerate(satislar, 2):
        worksheet.cell(row=row, column=1, value=satis.satis_no)
        worksheet.cell(
            row=row, column=2, value=satis.musteri.tam_ad if satis.musteri else 'Bilinmeyen')
        worksheet.cell(row=row, column=3, value=float(satis.toplam_tutar))
        worksheet.cell(row=row, column=4, value=satis.get_odeme_tipi_display())
        worksheet.cell(row=row, column=5,
                       value=satis.satis_tarihi.strftime('%d.%m.%Y %H:%M'))

    # Response
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="gunluk_satis_{secili_tarih}.xlsx"'
    workbook.save(response)
    return response


@login_required
def gunluk_satis_pdf(request):
    """Günlük satış PDF export"""
    # PDF oluşturma kodu buraya gelecek
    pass


@login_required
def stok_excel(request):
    """Stok raporu Excel export"""
    from urun.models import UrunVaryanti

    varyantlar = UrunVaryanti.objects.filter(
        aktif=True,
        urun__aktif=True
    ).select_related('urun', 'urun__kategori', 'urun__marka', 'renk', 'beden').order_by('urun__kategori__ad', 'urun__ad')

    # Filtreler - Güvenli şekilde
    try:
        arama = request.GET.get('arama', '').strip()
        kategori_id = request.GET.get('kategori')
        marka_id = request.GET.get('marka')
        durum = request.GET.get('durum')
        cinsiyet = request.GET.get('cinsiyet')
        kar_orani_min = request.GET.get('kar_orani_min', '').strip()
        kar_orani_max = request.GET.get('kar_orani_max', '').strip()

        # Arama filtresi
        if arama:
            from django.db.models import Q
            varyantlar = varyantlar.filter(
                Q(urun__ad__icontains=arama) |
                Q(barkod__icontains=arama) |
                Q(urun__urun_kodu__icontains=arama) |
                Q(renk__ad__icontains=arama) |
                Q(beden__ad__icontains=arama)
            )

        # Kategori filtresi - Güvenli None kontrolü
        if kategori_id and kategori_id != 'None' and kategori_id != '' and str(kategori_id).isdigit():
            varyantlar = varyantlar.filter(urun__kategori_id=int(kategori_id))

        # Marka filtresi - Güvenli None kontrolü
        if marka_id and marka_id != 'None' and marka_id != '' and str(marka_id).isdigit():
            varyantlar = varyantlar.filter(urun__marka_id=int(marka_id))

        # Stok durumu filtresi - Yeni 3 seçenek: Hepsi / Stoğu Olan / Stoğu Biten
        if durum and durum != 'None' and durum != '':
            if durum == 'stogu_olan':
                varyantlar = varyantlar.filter(stok_miktari__gt=0)
            elif durum == 'stogu_biten':
                varyantlar = varyantlar.filter(stok_miktari=0)
            # Boş veya yok → tümü

        # Cinsiyet filtresi - None kontrolü
        if cinsiyet and cinsiyet != 'None' and cinsiyet != '' and cinsiyet != 'hepsi':
            varyantlar = varyantlar.filter(urun__cinsiyet=cinsiyet)

        # Kar oranı filtre - Python tarafı
        varyant_list = list(varyantlar)
        min_val = None
        max_val = None
        try:
            if kar_orani_min:
                min_val = float(kar_orani_min)
        except (ValueError, TypeError):
            pass
        try:
            if kar_orani_max:
                max_val = float(kar_orani_max)
        except (ValueError, TypeError):
            pass

        if min_val is not None or max_val is not None:
            ZERO_D = Decimal('0')
            filtered = []
            for v in varyant_list:
                alis = v.urun.alis_fiyati or ZERO_D
                satis = v.urun.pesin_fiyat or ZERO_D
                if satis != 0:
                    oran = float((satis - alis) / satis * Decimal('100'))
                else:
                    oran = 0.0
                if min_val is not None and oran < min_val:
                    continue
                if max_val is not None and oran > max_val:
                    continue
                filtered.append(v)
            varyantlar = filtered
        else:
            varyantlar = varyant_list

    except (ValueError, TypeError) as e:
        # Hata durumunda filtreleri atla, tüm varyantları getir
        pass

    # Excel dosyası oluştur
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Stok Raporu"

    # Başlıklar
    headers = ['Ürün Adı', 'Varyant', 'Barkod', 'Kategori', 'Marka', 'Cinsiyet',
               'Alış Fiyatı', 'Satış Fiyatı', 'Kar Oranı %', 'Kar Tutarı', 'Stok Miktarı', 'Durum']
    for col, header in enumerate(headers, 1):
        worksheet.cell(row=1, column=col, value=header)

    # Veriler
    for row, varyant in enumerate(varyantlar, 2):
        varyant_adi = ""
        if varyant.renk:
            varyant_adi += varyant.renk.ad
        if varyant.renk and varyant.beden:
            varyant_adi += " - "
        if varyant.beden:
            varyant_adi += varyant.beden.ad
        if not varyant_adi:
            varyant_adi = "Standart"

        durum_text = "Normal"
        if varyant.stok_miktari == 0:
            durum_text = "Tükendi"
        elif varyant.stok_miktari <= 5:
            durum_text = "Kritik"

        worksheet.cell(row=row, column=1, value=varyant.urun.ad)
        worksheet.cell(row=row, column=2, value=varyant_adi)
        worksheet.cell(row=row, column=3, value=varyant.barkod)
        worksheet.cell(row=row, column=4, value=str(varyant.urun.kategori))
        worksheet.cell(row=row, column=5, value=str(
            varyant.urun.marka) if varyant.urun.marka else "-")
        worksheet.cell(row=row, column=6,
                       value=varyant.urun.get_cinsiyet_display())
        worksheet.cell(row=row, column=7, value=float(
            varyant.urun.alis_fiyati))
        worksheet.cell(row=row, column=8, value=float(
            varyant.urun.pesin_fiyat))  # 'satis_fiyati' yerine 'pesin_fiyat'
        # Kar oranı ve kâr tutarı dinamik hesaplama
        _alis = varyant.urun.alis_fiyati or Decimal('0')
        _satis = varyant.urun.pesin_fiyat or Decimal('0')
        _tutar = _satis - _alis
        if _satis != 0:
            _oran = float(_tutar / _satis * Decimal('100'))
        else:
            _oran = 0.0
        worksheet.cell(row=row, column=9, value=_oran)
        worksheet.cell(row=row, column=10, value=float(_tutar))
        worksheet.cell(row=row, column=11, value=varyant.stok_miktari)
        worksheet.cell(row=row, column=12, value=durum_text)

    # Response
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="stok_raporu.xlsx"'
    workbook.save(response)
    return response


@login_required
def stok_pdf(request):
    """Stok raporu PDF export"""
    # PDF oluşturma kodu buraya gelecek
    pass


@login_required
def kar_zarar_excel(request):
    """Kâr/Zarar Excel export"""
    # Excel oluşturma kodu buraya gelecek
    pass


@login_required
def kar_zarar_pdf(request):
    """Kâr/Zarar PDF export"""
    # PDF oluşturma kodu buraya gelecek
    pass


@login_required
def stok_hareketleri(request, varyant_id):
    """Ürün varyantının stok hareketleri"""
    from urun.models import UrunVaryanti, StokHareket
    from satis.models import SatisDetay
    from django.shortcuts import get_object_or_404

    varyant = get_object_or_404(UrunVaryanti, id=varyant_id)

    # Satış hareketleri (çıkışlar)
    satis_hareketleri = SatisDetay.objects.filter(
        varyant=varyant
    ).select_related('satis', 'satis__musteri', 'satis__satici').order_by('-satis__satis_tarihi')

    # Stok hareketleri (giriş, çıkış, düzeltme vb.)
    stok_hareketleri = StokHareket.objects.filter(
        varyant=varyant
    ).select_related('kullanici').order_by('-olusturma_tarihi')

    context = {
        'varyant': varyant,
        'satis_hareketleri': satis_hareketleri,
        'stok_hareketleri': stok_hareketleri,
        'title': f'{varyant.urun.ad} - Stok Hareketleri'
    }
    return render(request, 'rapor/stok_hareketleri.html', context)


@login_required
def satici_raporu(request):
    """Satış elemanlarının performans raporu"""
    from django.contrib.auth import get_user_model
    from django.db.models import Sum, Count, Avg
    from datetime import datetime, timedelta, date
    from decimal import Decimal

    User = get_user_model()

    # Tarih filtreleri
    bugun = date.today()
    filtre = request.GET.get('filtre', 'ay')  # gun, hafta, ay, ozel

    if filtre == 'gun':
        baslangic = bugun
        bitis = bugun
        baslik = f"Günlük Satış Raporu - {bugun.strftime('%d.%m.%Y')}"
    elif filtre == 'hafta':
        # Bu haftanın başından bugüne
        haftanin_basi = bugun - timedelta(days=bugun.weekday())
        baslangic = haftanin_basi
        bitis = bugun
        baslik = f"Haftalık Satış Raporu - {haftanin_basi.strftime('%d.%m')} / {bugun.strftime('%d.%m.%Y')}"
    elif filtre == 'ay':
        # Bu ayın başından bugüne
        ayin_basi = bugun.replace(day=1)
        baslangic = ayin_basi
        bitis = bugun
        baslik = f"Aylık Satış Raporu - {ayin_basi.strftime('%B %Y')}"
    else:  # ozel
        baslangic_str = request.GET.get(
            'baslangic', bugun.strftime('%Y-%m-%d'))
        bitis_str = request.GET.get('bitis', bugun.strftime('%Y-%m-%d'))

        try:
            baslangic = datetime.strptime(baslangic_str, '%Y-%m-%d').date()
            bitis = datetime.strptime(bitis_str, '%Y-%m-%d').date()
        except ValueError:
            baslangic = bugun
            bitis = bugun

        baslik = f"Özel Dönem Satış Raporu - {baslangic.strftime('%d.%m.%Y')} / {bitis.strftime('%d.%m.%Y')}"

    # Satış elemanlarını al (satış yapanları)
    satici_listesi = User.objects.filter(
        satis__satis_tarihi__date__range=[baslangic, bitis],
        satis__durum='tamamlandi'
    ).distinct()

    # Her satış elemanı için istatistikler (ciro: genel_toplam = indirim sonrası net tahsilat)
    satici_stats = []
    toplam_satis_tutari = Decimal('0')
    toplam_satis_adedi = 0
    toplam_urun_adedi = 0

    for satici in satici_listesi:
        # Bu satıcının dönemdeki satışları
        satislar = Satis.objects.filter(
            satici=satici,
            satis_tarihi__date__range=[baslangic, bitis],
            durum='tamamlandi'
        )

        # İstatistikler — toplam_tutar (KDV+ara) değil; müşterinin ödediği net: genel_toplam
        stats = satislar.aggregate(
            toplam_tutar=Sum('genel_toplam'),
            satis_sayisi=Count('id')
        )

        # Ortalama: net fatura başına
        ortalama_satis = satislar.aggregate(ortalama=Avg('genel_toplam'))[
            'ortalama'] or Decimal('0')

        # Satılan ürün sayısı
        urun_sayisi = SatisDetay.objects.filter(
            satis__satici=satici,
            satis__satis_tarihi__date__range=[baslangic, bitis],
            satis__durum='tamamlandi'
        ).aggregate(toplam_adet=Sum('miktar'))['toplam_adet'] or 0

        # En çok sattığı ürün (satır neti: toplam_fiyat indirim düşülmüş)
        en_cok_satan = SatisDetay.objects.filter(
            satis__satici=satici,
            satis__satis_tarihi__date__range=[baslangic, bitis],
            satis__durum='tamamlandi'
        ).values(
            'urun__ad'
        ).annotate(
            toplam_adet=Sum('miktar'),
            toplam_tutar=Sum('toplam_fiyat')
        ).order_by('-toplam_adet').first()

        net_toplam = stats['toplam_tutar'] or Decimal('0')
        if stats['satis_sayisi']:
            satici_stats.append({
                'satici': satici,
                'toplam_tutar': net_toplam,
                'satis_sayisi': stats['satis_sayisi'],
                'urun_sayisi': urun_sayisi,
                'ortalama_satis': ortalama_satis,
                'en_cok_satan': en_cok_satan,
                'yuzde_pay': Decimal('0')  # Sonra hesaplanacak
            })

            toplam_satis_tutari += net_toplam
            toplam_satis_adedi += stats['satis_sayisi']
            toplam_urun_adedi += urun_sayisi

    # Yüzde paylarını hesapla
    for stat in satici_stats:
        if toplam_satis_tutari > 0:
            stat['yuzde_pay'] = (stat['toplam_tutar'] /
                                 toplam_satis_tutari) * 100

    # Performansa göre sırala
    satici_stats.sort(key=lambda x: x['toplam_tutar'], reverse=True)

    # Günlük detay için (sadece son 7 gün)
    gunluk_detay = []
    if filtre in ['gun', 'hafta']:
        son_7_gun = []
        for i in range(7):
            tarih = bugun - timedelta(days=i)
            if tarih >= baslangic:
                son_7_gun.append(tarih)

        for tarih in reversed(son_7_gun):
            gun_satislari = Satis.objects.filter(
                satis_tarihi__date=tarih,
                durum='tamamlandi'
            ).aggregate(
                toplam=Sum('genel_toplam'),
                adet=Count('id')
            )

            gunluk_detay.append({
                'tarih': tarih,
                'toplam_tutar': gun_satislari['toplam'] or Decimal('0'),
                'satis_adet': gun_satislari['adet'] or 0
            })

    context = {
        'title': 'Satış Elemanı Raporu',
        'baslik': baslik,
        'filtre': filtre,
        'baslangic': baslangic,
        'bitis': bitis,
        'satici_stats': satici_stats,
        'toplam_satis_tutari': toplam_satis_tutari,
        'toplam_satis_adedi': toplam_satis_adedi,
        'toplam_urun_adedi': toplam_urun_adedi,
        'gunluk_detay': gunluk_detay,
        'en_basarili_satici': satici_stats[0] if satici_stats else None,
    }

    return render(request, 'rapor/satici_raporu.html', context)


@login_required
def satici_raporu_excel(request):
    """Satıcı raporu Excel export"""
    # Export mantığı için satici_raporu view'ini tekrar çağır
    response = satici_raporu(request)
    # Excel export implementasyonu burada olacak
    return HttpResponse("Excel export yakında eklenecek")


@login_required
def satici_raporu_pdf(request):
    """Satıcı raporu PDF export"""
    # Export mantığı için satici_raporu view'ini tekrar çağır
    response = satici_raporu(request)
    # PDF export implementasyonu burada olacak
    return HttpResponse("PDF export yakında eklenecek")
